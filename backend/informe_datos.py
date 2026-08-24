"""Genera el informe del pistoleo en el formato Punteo de App movil.xlsm.

Replica el layout de la hoja Punteo Inicial / Punteo Final (VIVEROS ELCHE,
GGN 8438002215009) leyendo BigQuery: picking_registros + GestionComercialVE
(PEDIDOS, CLIENTE, LINEA_PEDIDO, ARTICULOS, CODIGOS_EAN, LITRAJES, SECTORES)
+ matriculas_pedido.

El libro contiene una hoja "Punteo" por cada parte (I/F) del pedido, una hoja
"Detalle" con cada evento del pistoleo y una hoja "Control" con el acopio por
linea frente al pedido.
"""

import io

from google.cloud import bigquery
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

GGN = "8438002215009"
DOC_PUNTEO_INICIAL = "Punteo Inicial"
DOC_PUNTEO_FINAL = "Punteo Final"
TABLE_ROWS = 41          # filas de datos por tabla (17-57 / 81-121)
BLOCK_OFFSET = 64        # separacion entre bloque 1 (filas 1-64) y bloque 2 (65-128)


def _set(v):
    return v if v is not None else ""


def _num(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


# Columnas de la tabla de datos. Tabla 1 (A-S) y tabla 2 (U-AL) por pagina.
_T1 = {"seq": 1, "ref": 2, "equiv": 4, "desc": 5, "talla": 13, "sector": 15, "finca": 17, "cant": 19}
_T2 = {"seq": 20, "ref": 21, "equiv": 23, "desc": 24, "talla": 32, "sector": 34, "finca": 36, "cant": 38}


def _load_datos(
    bq_client,
    project: str,
    dataset: str,
    picking_dataset: str,
    picking_table: str,
    matriculas_table: str,
    numero_pedido: str,
) -> dict:
    """Carga todos los datos del informe Punteo desde BigQuery (compartido por xlsx y pdf)."""
    params = [bigquery.ScalarQueryParameter("pedido", "STRING", numero_pedido)]
    jc = bigquery.QueryJobConfig(query_parameters=params)

    # 1) Cabecera del pedido
    order_sql = f"""
        SELECT p.SERIE_PEDIDO, p.NUMERO_PEDIDO, p.NUMERO_CLIENTE, p.FECHA_CARGA, p.SECTOR_CARGA, p.FINCA_CARGA,
               COALESCE(p.OBSERVACIONES, '') AS OBSERVACIONES,
               COALESCE(p.MARCA_PEDIDO, '') AS MARCA_PEDIDO,
               COALESCE(p.REFERENCIA_PEDIDO, '') AS REFERENCIA_PEDIDO,
               COALESCE(c.N_COMERCIAL, '') AS N_COMERCIAL, COALESCE(c.N_FISCAL, '') AS N_FISCAL,
               COALESCE(c.DIRECCION, '') AS DIR_CLIENTE,
               COALESCE(c.CIUDAD, '') AS CIUDAD_CLIENTE,
               COALESCE(mc.matricula, '') AS MATRICULA_CAMION,
               COALESCE(mr.matricula, '') AS MATRICULA_REMOLQUE
        FROM `{project}.{dataset}.PEDIDOS` p
        LEFT JOIN `{project}.{dataset}.CLIENTE` c ON c.ID_CLIENTE = p.NUMERO_CLIENTE
        LEFT JOIN `{project}.{picking_dataset}.{matriculas_table}` mc
            ON mc.pedido_id = p.NUMERO_PEDIDO AND mc.tipo = 'CAMION'
        LEFT JOIN `{project}.{picking_dataset}.{matriculas_table}` mr
            ON mr.pedido_id = p.NUMERO_PEDIDO AND mr.tipo = 'REMOLQUE'
        WHERE p.NUMERO_PEDIDO = @pedido
        LIMIT 1
    """
    order_rows = [dict(r) for r in bq_client.query(order_sql, job_config=jc).result()]
    if not order_rows:
        raise ValueError("Pedido no encontrado")
    o = order_rows[0]

    # 4) Detalle de cada evento del pistoleo (primero: alimenta partes y sustituciones)
    detalle_sql = f"""
        SELECT r.picking_numero, r.picking_tipo, r.fecha_hora, r.empleado_nombre,
               r.ean_escaneado, r.ocr_texto, r.ref_original, r.ref_servida,
               r.sustituido, r.litros, r.medida, r.calibre, r.cantidad_partida,
               l.POSICION_PEDIDO AS POSICION, l.REFERENCIA_ARTICULO AS REF_LINEA,
               a.DESCRIPCION_ARTICULO AS DESC_SERVIDA,
               lit.DESCRIPCION_LITRAJE AS TALLA, sec.DESCRIPCION_SECTOR AS SECTOR,
               litp.DESCRIPCION_LITRAJE AS LITRAJE_PEDIDA, secp.DESCRIPCION_SECTOR AS SECTOR_PEDIDA,
               lits.DESCRIPCION_LITRAJE AS LITRAJE_SERVIDA, secs.DESCRIPCION_SECTOR AS SECTOR_SERVIDA
        FROM `{project}.{picking_dataset}.{picking_table}` r
        LEFT JOIN `{project}.{dataset}.LINEA_PEDIDO` l ON l.HUELLA_DIGITAL = r.order_line_id
        LEFT JOIN `{project}.{dataset}.CODIGOS_EAN` ce ON ce.CODIGO_EAN = r.ean_escaneado
        LEFT JOIN `{project}.{dataset}.LITRAJES` lit
            ON lit.ID_LITRAJE = COALESCE(ce.CODIGO_LITRAJE, l.CODIGO_LITRAJE)
        LEFT JOIN `{project}.{dataset}.SECTORES` sec
            ON sec.ID_SECTOR = COALESCE(ce.CODIGO_SECTOR, l.CODIGO_SECTOR)
        LEFT JOIN `{project}.{dataset}.LITRAJES` litp ON litp.ID_LITRAJE = l.CODIGO_LITRAJE
        LEFT JOIN `{project}.{dataset}.SECTORES` secp ON secp.ID_SECTOR = l.CODIGO_SECTOR
        LEFT JOIN `{project}.{dataset}.LITRAJES` lits ON lits.ID_LITRAJE = ce.CODIGO_LITRAJE
        LEFT JOIN `{project}.{dataset}.SECTORES` secs ON secs.ID_SECTOR = ce.CODIGO_SECTOR
        LEFT JOIN `{project}.{dataset}.ARTICULOS` a ON a.ID_ARTICULO = r.ref_servida
        WHERE r.order_id = @pedido
        ORDER BY r.fecha_hora
    """
    detalle = [dict(r) for r in bq_client.query(detalle_sql, job_config=jc).result()]

    # 2) Partes (sesiones de pistoleo) del pedido
    partes_sql = f"""
        SELECT picking_numero, picking_tipo,
               MIN(fecha_hora) AS INICIO, MAX(fecha_hora) AS FIN,
               COUNT(*) AS EVENTOS
        FROM `{project}.{picking_dataset}.{picking_table}`
        WHERE order_id = @pedido
        GROUP BY picking_numero, picking_tipo
        ORDER BY picking_numero, picking_tipo
    """
    partes = [dict(r) for r in bq_client.query(partes_sql, job_config=jc).result()]
    empleados_por_parte: dict[tuple, set] = {}
    for r in detalle:
        key = (r.get("picking_numero"), r.get("picking_tipo"))
        name = r.get("empleado_nombre")
        if name:
            empleados_por_parte.setdefault(key, set()).add(str(name))
    for p in partes:
        names = empleados_por_parte.get((p["picking_numero"], p["picking_tipo"]), set())
        p["EMPLEADO"] = ", ".join(sorted(names))

    # 3) Filas del Punteo (agrupadas por parte + referencia servida)
    filas_sql = f"""
        SELECT r.picking_numero, r.picking_tipo, r.ref_servida,
               ANY_VALUE(l.POSICION_PEDIDO) AS POSICION,
               ANY_VALUE(a.DESCRIPCION_ARTICULO) AS DESCRIPCION,
               ANY_VALUE(a.GLOBALGAP) AS EQUIVALENTE,
               ANY_VALUE(a.FINCA_ARTICULO) AS FINCA_ARTICULO,
               ANY_VALUE(lit.DESCRIPCION_LITRAJE) AS TALLA,
               ANY_VALUE(sec.DESCRIPCION_SECTOR) AS SECTOR,
               SUM(r.cantidad_partida) AS CANT,
               COUNTIF(r.sustituido) AS SUSTITUCIONES
        FROM `{project}.{picking_dataset}.{picking_table}` r
        LEFT JOIN `{project}.{dataset}.ARTICULOS` a ON a.ID_ARTICULO = r.ref_servida
        LEFT JOIN `{project}.{dataset}.CODIGOS_EAN` ce ON ce.CODIGO_EAN = r.ean_escaneado
        LEFT JOIN `{project}.{dataset}.LINEA_PEDIDO` l ON l.HUELLA_DIGITAL = r.order_line_id
        LEFT JOIN `{project}.{dataset}.LITRAJES` lit
            ON lit.ID_LITRAJE = COALESCE(ce.CODIGO_LITRAJE, l.CODIGO_LITRAJE)
        LEFT JOIN `{project}.{dataset}.SECTORES` sec
            ON sec.ID_SECTOR = COALESCE(ce.CODIGO_SECTOR, l.CODIGO_SECTOR)
        WHERE r.order_id = @pedido
        GROUP BY r.picking_numero, r.picking_tipo, r.ref_servida
        ORDER BY r.picking_numero, r.picking_tipo, r.ref_servida
    """
    filas = [dict(r) for r in bq_client.query(filas_sql, job_config=jc).result()]

    # Sustituciones por (parte, ref_servida) para mostrar "orig -> serv"
    sust_map: dict[tuple, list[str]] = {}
    for r in detalle:
        if not r.get("sustituido"):
            continue
        key = (r.get("picking_numero"), r.get("picking_tipo"), r.get("ref_servida"))
        sust_map.setdefault(key, []).append(
            f"{r.get('ref_original') or ''} -> {r.get('ref_servida') or ''}"
        )

    # 5) Control de acopio por linea
    control_sql = f"""
        SELECT l.HUELLA_DIGITAL, l.POSICION_PEDIDO, l.REFERENCIA_ARTICULO,
               l.DESCRIPCION_ARTICULO, l.UNIDADES, l.UNIDADES_PENDIENTES,
               COALESCE(lit.DESCRIPCION_LITRAJE, '') AS TALLA,
               COALESCE(sec.DESCRIPCION_SECTOR, '') AS SECTOR,
               COALESCE(pr.ACOPIADO, 0) AS ACOPIADO,
               COALESCE(pr.SUSTITUIDO, FALSE) AS SUSTITUIDO,
               COALESCE(pr.PARTES, '') AS PARTES
        FROM `{project}.{dataset}.LINEA_PEDIDO` l
        LEFT JOIN `{project}.{dataset}.LITRAJES` lit ON lit.ID_LITRAJE = l.CODIGO_LITRAJE
        LEFT JOIN `{project}.{dataset}.SECTORES` sec ON sec.ID_SECTOR = l.CODIGO_SECTOR
        LEFT JOIN (
            SELECT order_line_id, SUM(cantidad_partida) AS ACOPIADO,
                   COUNTIF(sustituido) > 0 AS SUSTITUIDO,
                   STRING_AGG(DISTINCT CONCAT(CAST(picking_numero AS STRING), picking_tipo), ', ')
                       AS PARTES
            FROM `{project}.{picking_dataset}.{picking_table}`
            WHERE order_id = @pedido
            GROUP BY order_line_id
        ) pr ON pr.order_line_id = l.HUELLA_DIGITAL
        WHERE l.NUMERO_PEDIDO = @pedido AND COALESCE(l.IMPRIMIR_LINEA, 0) = 0
        ORDER BY l.POSICION_PEDIDO
    """
    control = [dict(r) for r in bq_client.query(control_sql, job_config=jc).result()]

    # 6) Referencias servidas sin localizar en el catalogo actual
    sin_localizar_sql = f"""
        SELECT DISTINCT r.ref_servida, r.ean_escaneado
        FROM `{project}.{picking_dataset}.{picking_table}` r
        LEFT JOIN `{project}.{dataset}.ARTICULOS` a ON a.ID_ARTICULO = r.ref_servida
        WHERE r.order_id = @pedido AND a.ID_ARTICULO IS NULL
        ORDER BY r.ref_servida
    """
    sin_localizar = [dict(r) for r in bq_client.query(sin_localizar_sql, job_config=jc).result()]

    return {
        "o": o,
        "partes": partes,
        "filas": filas,
        "detalle": detalle,
        "control": control,
        "sin_localizar": sin_localizar,
        "sust_map": sust_map,
    }


def build_punteo_xlsx(
    bq_client,
    project: str,
    dataset: str,
    picking_dataset: str,
    picking_table: str,
    matriculas_table: str,
    numero_pedido: str,
) -> bytes:
    """Devuelve los bytes del .xlsx del informe Punteo del pedido."""
    datos = _load_datos(
        bq_client, project, dataset, picking_dataset, picking_table,
        matriculas_table, numero_pedido,
    )
    return _build_workbook(
        pedido=numero_pedido,
        o=datos["o"],
        partes=datos["partes"],
        filas=datos["filas"],
        detalle=datos["detalle"],
        control=datos["control"],
        sin_localizar=datos["sin_localizar"],
    )


# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------

_FONT_TITLE = Font(name="Arial", size=14, bold=True)
_FONT_BOLD = Font(name="Arial", size=10, bold=True)
_FONT_NORMAL = Font(name="Arial", size=10)
_FONT_SMALL = Font(name="Arial", size=8)
_FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_FILL_HEADER = PatternFill("solid", fgColor="4472C4")
_FILL_SUBHEADER = PatternFill("solid", fgColor="DDEBF7")
_THIN = Side(style="thin", color="999999")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _style_cell(ws, coord, value, font=None, fill=None, border=False, align=None, number=None):
    c = ws[coord]
    c.value = value
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if border:
        c.border = _BORDER
    if align:
        c.alignment = align
    if number is not None:
        c.number_format = number
    return c


def _write_header_block(ws, block, o, documento):
    """Escribe el bloque de cabecera del Punteo. block 0 = filas 1-64, block 1 = 65-128."""
    off = BLOCK_OFFSET * block
    row = lambda r: r + off

    _style_cell(ws, f"A{row(1)}", "VIVEROS ELCHE, S.L.", _FONT_TITLE)
    _style_cell(ws, f"T{row(1)}", "VIVEROS ELCHE, S.L.", _FONT_TITLE)
    ws.merge_cells(f"A{row(1)}:D{row(1)}")
    ws.merge_cells(f"T{row(1)}:W{row(1)}")

    addr = "Ctra. de la Fábrica, s/n" if block == 0 else "PARTIDA DE ALGORÍS, POLÍGONO 1-189-A"
    _style_cell(ws, f"A{row(2)}", addr, _FONT_SMALL)
    _style_cell(ws, f"T{row(2)}", addr, _FONT_SMALL)
    _style_cell(ws, f"K{row(2)}", "FINCA", _FONT_BOLD, border=True, align=_CENTER)
    _style_cell(ws, f"M{row(2)}", _set(o.get("FINCA_CARGA")), _FONT_BOLD, border=True, align=_CENTER)
    _style_cell(ws, f"K{row(4)}", "ZONA", _FONT_BOLD, border=True, align=_CENTER)
    _style_cell(ws, f"M{row(4)}", _set(o.get("SECTOR_CARGA")), _FONT_BOLD, border=True, align=_CENTER)
    _style_cell(ws, f"AD{row(2)}", "FINCA", _FONT_BOLD, border=True, align=_CENTER)
    _style_cell(ws, f"AF{row(2)}", _set(o.get("FINCA_CARGA")), _FONT_BOLD, border=True, align=_CENTER)
    _style_cell(ws, f"AD{row(4)}", "ZONA", _FONT_BOLD, border=True, align=_CENTER)
    _style_cell(ws, f"AF{row(4)}", _set(o.get("SECTOR_CARGA")), _FONT_BOLD, border=True, align=_CENTER)

    _style_cell(ws, f"D{row(8)}", _set(o.get("N_COMERCIAL")), _FONT_NORMAL)
    _style_cell(ws, f"D{row(9)}", _set(o.get("N_FISCAL")), _FONT_NORMAL)
    _style_cell(ws, f"W{row(8)}", _set(o.get("N_COMERCIAL")), _FONT_NORMAL)
    _style_cell(ws, f"W{row(9)}", _set(o.get("N_FISCAL")), _FONT_NORMAL)
    ws.merge_cells(f"D{row(8)}:E{row(8)}")
    ws.merge_cells(f"D{row(9)}:E{row(9)}")
    ws.merge_cells(f"W{row(8)}:X{row(8)}")
    ws.merge_cells(f"W{row(9)}:X{row(9)}")

    hdr1 = {"A": "TRACTORA", "G": "REMOLQUE", "L": "DOCUMENTO", "N": "NÚMERO", "P": "PÁGINA", "R": "FECHA"}
    hdr2 = {"T": "TRACTORA", "Z": "REMOLQUE", "AD": "DOCUMENTO", "AF": "NÚMERO", "AH": "PÁGINA", "AJ": "FECHA"}
    for col, txt in hdr1.items():
        _style_cell(ws, f"{col}{row(12)}", txt, _FONT_BOLD, _FILL_SUBHEADER, True, _CENTER)
    for col, txt in hdr2.items():
        _style_cell(ws, f"{col}{row(12)}", txt, _FONT_BOLD, _FILL_SUBHEADER, True, _CENTER)

    _style_cell(ws, f"A{row(13)}", _set(o.get("MATRICULA_CAMION")), _FONT_NORMAL, border=True, align=_CENTER)
    _style_cell(ws, f"G{row(13)}", _set(o.get("MATRICULA_REMOLQUE")), _FONT_NORMAL, border=True, align=_CENTER)
    _style_cell(ws, f"L{row(13)}", documento, _FONT_NORMAL, border=True, align=_CENTER)
    _style_cell(ws, f"N{row(13)}", _set(o.get("NUMERO_PEDIDO")), _FONT_NORMAL, border=True, align=_CENTER)
    _style_cell(ws, f"P{row(13)}", "1 de 1", _FONT_NORMAL, border=True, align=_CENTER)
    fecha = str(o.get("FECHA_CARGA") or "")[:10]
    _style_cell(ws, f"R{row(13)}", fecha, _FONT_NORMAL, border=True, align=_CENTER)

    _style_cell(ws, f"T{row(13)}", _set(o.get("MATRICULA_CAMION")), _FONT_NORMAL, border=True, align=_CENTER)
    _style_cell(ws, f"Z{row(13)}", _set(o.get("MATRICULA_REMOLQUE")), _FONT_NORMAL, border=True, align=_CENTER)
    _style_cell(ws, f"AD{row(13)}", documento, _FONT_NORMAL, border=True, align=_CENTER)
    _style_cell(ws, f"AF{row(13)}", _set(o.get("NUMERO_PEDIDO")), _FONT_NORMAL, border=True, align=_CENTER)
    _style_cell(ws, f"AH{row(13)}", "1 de 1", _FONT_NORMAL, border=True, align=_CENTER)
    _style_cell(ws, f"AJ{row(13)}", fecha, _FONT_NORMAL, border=True, align=_CENTER)

    t1 = {"B": "REFERENCIA", "E": "DESCRIPCIÓN", "M": "TALLA", "O": "SECTOR", "Q": "FINCA", "S": "CARG."}
    t2 = {"U": "REFERENCIA", "X": "DESCRIPCIÓN", "AF": "TALLA", "AH": "SECTOR", "AJ": "FINCA", "AL": "CARG."}
    for col, txt in t1.items():
        _style_cell(ws, f"{col}{row(15)}", txt, _FONT_HEADER, _FILL_HEADER, True, _CENTER)
    for col, txt in t2.items():
        _style_cell(ws, f"{col}{row(15)}", txt, _FONT_HEADER, _FILL_HEADER, True, _CENTER)


def _write_footer_block(ws, block, peso, obs, empleado):
    """Escribe el pie del Punteo (filas 59-64 o 123-128)."""
    off = BLOCK_OFFSET * block
    row = lambda r: r + off

    _style_cell(ws, f"A{row(59)}", f"* Producto certificado GlobalG.A.P. GGN {GGN}", _FONT_SMALL)
    _style_cell(ws, f"T{row(59)}", f"* Producto certificado GlobalG.A.P. GGN {GGN}", _FONT_SMALL)
    _style_cell(ws, f"N{row(59)}", "Verificado por:", _FONT_SMALL)
    _style_cell(ws, f"AG{row(59)}", "Verificado por:", _FONT_SMALL)

    _style_cell(ws, f"A{row(61)}", "OBSERVACIONES:", _FONT_BOLD, border=True)
    _style_cell(ws, f"T{row(61)}", "OBSERVACIONES:", _FONT_BOLD, border=True)
    s_first = row(17)
    s_last = row(17) + TABLE_ROWS - 1
    al_first = row(17)
    al_last = row(17) + TABLE_ROWS - 1
    if block == 0:
        l61 = f'=IF(S{s_first}<>"",SUM(S{s_first}:S{s_last}),"")'
        al61 = f'=IF(AL{al_first}<>"",SUM(AL{al_first}:AL{al_last},L61),"")'
    else:
        l61 = f'=IF(S{s_first}<>"",SUM(AE61,L61,S{s_first}:S{s_last}),"")'
        al61 = f'=IF(AL{al_first}<>"",SUM(L{row(61)},AE61,L61,AL{al_first}:AL{al_last}),"")'
    _style_cell(ws, f"L{row(61)}", l61, _FONT_NORMAL, border=True, align=_CENTER)
    _style_cell(ws, f"AL{row(61)}", al61, _FONT_NORMAL, border=True, align=_CENTER)
    _style_cell(ws, f"N{row(61)}", _set(empleado), _FONT_NORMAL, border=True, align=_CENTER)
    _style_cell(ws, f"AG{row(61)}", _set(empleado), _FONT_NORMAL, border=True, align=_CENTER)
    _style_cell(ws, f"A{row(62)}", _set(obs), _FONT_NORMAL, border=True)
    _style_cell(ws, f"T{row(62)}", _set(obs), _FONT_NORMAL, border=True)

    page_left = block * 2 + 1
    page_right = block * 2 + 2
    _style_cell(ws, f"A{row(64)}", f"Página {page_left}", _FONT_NORMAL)
    _style_cell(ws, f"E{row(64)}", "PESO DE LA CARGA (KG)", _FONT_BOLD)
    _style_cell(ws, f"J{row(64)}", _num(peso), _FONT_NORMAL, border=True, align=_CENTER, number="0.00")
    _style_cell(ws, f"T{row(64)}", f"Página {page_right}", _FONT_NORMAL)
    _style_cell(ws, f"X{row(64)}", "PESO DE LA CARGA (KG)", _FONT_BOLD)
    _style_cell(ws, f"AC{row(64)}", _num(peso), _FONT_NORMAL, border=True, align=_CENTER, number="0.00")


def _write_table_row(ws, row_idx, seq, ref, equiv, desc, talla, sector, finca, cant, table):
    """Escribe una fila de datos del Punteo. table = _T1 (A-S) o _T2 (U-AL)."""
    def col(key):
        return get_column_letter(table[key])

    _style_cell(ws, f"{col('seq')}{row_idx}", seq, _FONT_NORMAL, border=True, align=_CENTER)
    _style_cell(ws, f"{col('ref')}{row_idx}", _set(ref), _FONT_BOLD, border=True, align=_CENTER)
    _style_cell(ws, f"{col('equiv')}{row_idx}", _set(equiv), _FONT_NORMAL, border=True)
    _style_cell(ws, f"{col('desc')}{row_idx}", _set(desc), _FONT_NORMAL, border=True)
    _style_cell(ws, f"{col('talla')}{row_idx}", _set(talla), _FONT_NORMAL, border=True, align=_CENTER)
    _style_cell(ws, f"{col('sector')}{row_idx}", _set(sector), _FONT_NORMAL, border=True, align=_CENTER)
    _style_cell(ws, f"{col('finca')}{row_idx}", _set(finca), _FONT_NORMAL, border=True, align=_CENTER)
    _style_cell(ws, f"{col('cant')}{row_idx}", _num(cant), _FONT_NORMAL, border=True, align=_CENTER, number="0")


def _build_workbook(pedido, o, partes, filas, detalle, control, sin_localizar):
    wb = Workbook()
    ws0 = wb.create_sheet("Portada", 0)
    wb.remove(wb["Sheet"])

    obs = _set(o.get("OBSERVACIONES"))
    peso = 0.0  # pendiente de capturar en el cierre de parte (D-153)

    # --- Hojas Punteo (una por parte) -------------------------------------
    for parte in partes:
        pnum = parte["picking_numero"]
        ptype = parte["picking_tipo"]
        documento = DOC_PUNTEO_INICIAL if ptype == "I" else DOC_PUNTEO_FINAL
        empleado = _set(parte.get("EMPLEADO"))
        rows = [f for f in filas if f["picking_numero"] == pnum and f["picking_tipo"] == ptype]
        if not rows:
            continue

        ws = wb.create_sheet(f"Punteo {ptype} {pnum}")

        for block in range(2):
            _write_header_block(ws, block, o, documento)
            _write_footer_block(ws, block, peso, obs, empleado)
            off = BLOCK_OFFSET * block
            chunk = rows[block * 2 * TABLE_ROWS: (block + 1) * 2 * TABLE_ROWS]
            data_start = 17 + off
            n1 = min(len(chunk), TABLE_ROWS)
            for i, r in enumerate(chunk[:n1]):
                _write_table_row(ws, data_start + i, i + 1, r.get("ref_servida"),
                                 r.get("EQUIVALENTE"), r.get("DESCRIPCION"), r.get("TALLA"),
                                 r.get("SECTOR"), r.get("FINCA_ARTICULO"), r.get("CANT"), _T1)
            for i, r in enumerate(chunk[n1:]):
                _write_table_row(ws, data_start + i, i + 1, r.get("ref_servida"),
                                 r.get("EQUIVALENTE"), r.get("DESCRIPCION"), r.get("TALLA"),
                                 r.get("SECTOR"), r.get("FINCA_ARTICULO"), r.get("CANT"), _T2)
            if len(rows) <= (block + 1) * 2 * TABLE_ROWS:
                break

    # --- Hoja Detalle -------------------------------------------------------
    ws = wb.create_sheet("Detalle")
    headers = ["Parte", "Tipo", "Fecha / Hora", "Empleado", "Pos", "Ref. Pedida (línea)",
               "Ref. Servida", "Cantidad", "Sustituido", "EAN", "OCR / Pasaporte",
               "Litros", "Medida", "Calibre", "Talla", "Sector"]
    for c, h in enumerate(headers, start=1):
        _style_cell(ws, f"{get_column_letter(c)}1", h, _FONT_HEADER, _FILL_HEADER, True, _CENTER)
    for i, r in enumerate(detalle, start=2):
        vals = [
            r.get("picking_numero"), r.get("picking_tipo"),
            str(r.get("fecha_hora") or "").replace("T", " ")[:19],
            r.get("empleado_nombre"), r.get("POSICION"), r.get("ref_original"),
            r.get("ref_servida"), _num(r.get("cantidad_partida")),
            "SÍ" if r.get("sustituido") else "",
            r.get("ean_escaneado"), r.get("ocr_texto"),
            _num(r.get("litros")), r.get("medida"), r.get("calibre"),
            r.get("TALLA"), r.get("SECTOR"),
        ]
        for c, v in enumerate(vals, start=1):
            col = get_column_letter(c)
            num = "#0" if c in (1, 6, 7, 8) else None
            _style_cell(ws, f"{col}{i}", v, _FONT_NORMAL, border=True,
                        align=_LEFT if c not in (1, 3, 4, 8, 9) else _CENTER, number=num)
    widths = [6, 6, 19, 24, 5, 14, 14, 9, 10, 15, 16, 8, 10, 9, 12, 12]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(detalle) + 1)}"

    # --- Hoja Control --------------------------------------------------------
    ws = wb.create_sheet("Control")
    headers = ["Pos", "Ref. Pedida", "Descripción", "Talla", "Sector", "Pedido",
               "Pendiente", "Pistoleado", "Dif. (Pendiente)", "Sustituido", "Partes"]
    for c, h in enumerate(headers, start=1):
        _style_cell(ws, f"{get_column_letter(c)}1", h, _FONT_HEADER, _FILL_HEADER, True, _CENTER)
    r_i = 2
    for l in control:
        pend = _num(l.get("UNIDADES_PENDIENTES"))
        pist = _num(l.get("ACOPIADO"))
        dif = pend - max(pist, 0)
        vals = [l.get("POSICION_PEDIDO"), l.get("REFERENCIA_ARTICULO"),
                l.get("DESCRIPCION_ARTICULO"), l.get("TALLA"), l.get("SECTOR"),
                _num(l.get("UNIDADES")), pend, pist, dif,
                "SÍ" if l.get("SUSTITUIDO") else "", l.get("PARTES")]
        for c, v in enumerate(vals, start=1):
            col = get_column_letter(c)
            _style_cell(ws, f"{col}{r_i}", v, _FONT_NORMAL, border=True,
                        align=_LEFT if c in (2, 3) else _CENTER)
        r_i += 1
    if sin_localizar:
        r_i += 1
        ws.merge_cells(start_row=r_i, start_column=1, end_row=r_i, end_column=11)
        _style_cell(ws, f"A{r_i}", "REFERENCIAS SERVIDAS NO LOCALIZADAS EN EL CATÁLOGO 2026 (revisar)",
                    _FONT_BOLD, _FILL_SUBHEADER, True, _LEFT)
        r_i += 1
        for s in sin_localizar:
            _style_cell(ws, f"A{r_i}", _set(s.get("ref_servida")), _FONT_NORMAL, border=True, align=_CENTER)
            _style_cell(ws, f"B{r_i}", _set(s.get("ean_escaneado")), _FONT_NORMAL, border=True, align=_CENTER)
            r_i += 1
    widths = [5, 14, 34, 12, 10, 8, 10, 11, 12, 11, 16]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    # Anchuras de las hojas Punteo
    for ws in wb.worksheets:
        if ws.title.startswith("Punteo"):
            for col in "ABCDEFGHIJKLMNOPQRSTUVWXYZABCDEFGHIJKLMN":
                ws.column_dimensions[col].width = 10
            ws.column_dimensions["E"].width = 30
            ws.column_dimensions["X"].width = 30

    # Portada con resumen
    ws0["A1"] = "INFORME PUNTEO - PEDIDO " + str(pedido)
    ws0["A1"].font = Font(name="Arial", size=16, bold=True)
    ws0["A2"] = _set(o.get("N_COMERCIAL"))
    finca = _set(o.get("FINCA_CARGA"))
    zona = _set(o.get("SECTOR_CARGA"))
    fecha_carga = str(o.get("FECHA_CARGA") or "")[:10]
    ws0["A4"] = f"Finca: {finca}   Zona: {zona}   Fecha carga: {fecha_carga}"
    partes_txt = ", ".join(f"{p['picking_tipo']} {p['picking_numero']}" for p in partes) or "Sin partes"
    ws0["A6"] = "Partes: " + partes_txt
    ws0.column_dimensions["A"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()